from django.db.models import Max, Count

from .serializers import *
from .models import *
from .services import audio, video, session
import uuid
from datetime import datetime, timedelta
from semantic_text_similarity.models import WebBertSimilarity

from django.shortcuts import get_object_or_404, get_list_or_404
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

web_model = WebBertSimilarity(device='cpu', batch_size=10)
from django.contrib.auth.hashers import make_password
from pandas import *
import os
import openpyxl
import random
from django.http import HttpResponse
import zipfile
from .services import text_service

class TextTestView(APIView):
    #   permission_classes = (IsAuthenticated,)
    def get(self, request, session_id):

        session = get_object_or_404(Session, id=session_id)
        text_analysis_results = text_service.analyzeTextIt(session.text)
        print(text_analysis_results)

        return Response({
            "result": text_analysis_results
        })

#---------------------------BACKOFFICE-------------------------------------
class EmployeeStatsAPIView(APIView):
    #   permission_classes = (IsAuthenticated,)
    def get(self, request):
        companyId = request.session['companyId']
        numEmployees = Employee.objects.filter(company=companyId).count()
        activeEmployees = Employee.objects.filter(id__in=Session.objects.values_list('employee_id', flat=True)).count()

        return Response({
            "numEmployees": numEmployees,
            "activeEmployees": activeEmployees
        })

class NewEmployee(APIView):
    def post(self, request):


        user = AppUsers.objects.create(email=request.data['email'], password=request.data['password'], is_active=True,
                                       is_staff=False,
                                       is_superuser=False)
        data = {}
        data['user'] = user.id
        data['step'] = 0
        data['session_no'] = 0
        data['company'] = 1
        data['name'] = request.data['name']
        data['surname'] = request.data['surname']

        serializer = EmployeeUserStepSerializer(data=data)
        if serializer.is_valid():
            user.set_password(request.data['password'])
            user.save()
            serializer.save()

            return Response(status=status.HTTP_201_CREATED)

        return Response(data=serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class StressStats(APIView):
    def get(self, request):
        stats = Employee.objects.annotate(sessions=Count('session__id'))
        print(stats)

        return Response(stats)

class StressStatsTimespan(APIView):
    def get(self, request, timespan):

        today = datetime.today()

        if timespan == "week":
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=7)

            stats = Session.objects.filter(date__gte=week_start, date__lt=week_end)
        elif timespan == "month":
            stats = Session.objects.filter(date__month=today.month, date__year=today.year)
        elif timespan == "year":
            stats = Session.objects.filter(date__year=today.year)
        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)

        stats = stats.order_by('date')

        serializer = SessionSerializer(stats, many=True)

        return Response(serializer.data)

class GetInteractionSummary(APIView):
    def get(self, request):
        company_id = request.session['companyId']

        # query = "SELECT id, name, surname, is_stressed, max(date) FROM employees natural join chatsession WHERE company = %s GROUP BY id, name, surname, is_stressed"

        result = Session.objects \
            .filter(employee__company=company_id) \
            .values('employee__id', 'employee__username', 'employee__stressed', 'employee__session_no') \
            .annotate(lastDate=Max('date')) \
            .order_by('employee__id')

        return Response(result)

class GetUserInteractions(APIView):
    def get(self, request, employee_id):
        # query = "SELECT id, date, first_prevailing_emotion FROM chatsession WHERE employee=%d ORDER BY date DESC"

        result = Session.objects.filter(employee=employee_id, analyzed=True).order_by("-date")
        serializer = SessionSerializer(result, many=True).data

        for s in serializer:
            firstEmotion = get_object_or_404(Emotion, id=s['first_prevailing_emotion'])
            s['first_prevailing_emotion'] = EmotionsSerializer(firstEmotion).data['extended_name']

        return Response(serializer)

class EmployeeAPIView(APIView):

    def get(self, request):
        employee = Employee.objects.all()
        item = self.request.query_params.get("item", "")

        if item != "":
            employee = employee.filter(name=item)

        serializer = EmployeeSerializer(employee, many=True)

        return Response(serializer.data)

    def post(self, request):
        serializer = EmployeeSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(status=status.HTTP_201_CREATED)

        return Response(data=serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class EmployeeDetailAPIView(APIView):

    def get(self, request, employee_id):
        employee = get_object_or_404(Employee, id=employee_id)
        serializer = EmployeeSerializer(employee)
        return Response(data=serializer.data, status=status.HTTP_200_OK)

class RetrieveEmployeeInformation(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        user_id = request.data.get('user_id', None)
        if user_id is not None:
            emp = EmployeeDataSerializer(get_object_or_404(Employee, user=get_object_or_404(AppUsers, pk=user_id))).data
            return Response({"employee": emp})
        else:
            return Response("User id not found")

class RetrieveEmployerInformation(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        user_id = request.data.get('user_id', None)
        if user_id is not None:
            emp = EmployerSerializer(get_object_or_404(Employer, user=get_object_or_404(AppUsers, pk=user_id))).data
            request.session['companyId'] = emp['company']

            return Response({"employer": emp})
        else:
            return Response("User id not found")

class RetrieveSessionsEmployee(APIView):
    def get(self, request, employee_id):
        chats = SessionMiniSerializer(get_list_or_404(Session.objects.filter(employee=get_object_or_404(Employee, pk=employee_id), analyzed=True).order_by("-date")), many=True).data
        return Response({"chats": chats})

class InteractionDetailsAPIView(APIView):
    def get(self, request, session_id):
        session = get_object_or_404(Session, id=session_id)
        req = get_object_or_404(Request, id=session.request_id)
        results = SessionResults.objects.filter(session_id=session_id)
        results = ResultsSerializerWithSession(results, many=True)
        results = results.data[:]
        response = {}
        for r in results:
            if(r['text'] == True):
                response['text_results'] = r
            if(r['audio'] == True):
                response['audio_results'] = r
            if(r['video'] == True):
                response['video_results'] = r
        csv_results = None
        serializer = SessionSerializer(session).data
        if(serializer['first_prevailing_emotion']):
            firstEmotion = get_object_or_404(Emotion, id=serializer['first_prevailing_emotion'])
            serializer['first_prevailing_emotion'] = EmotionsSerializer(firstEmotion).data['extended_name']
        if(serializer['second_prevailing_emotion']):
            secondEmotion = get_object_or_404(Emotion, id=serializer['second_prevailing_emotion'])
            serializer['second_prevailing_emotion'] = EmotionsSerializer(secondEmotion).data['extended_name']
        if(serializer['full_video_path']):
            serializer['hasGraph'] = True
            data = read_csv("tmp/{}/video_analysis.csv".format(session_id))

            hp = data['hp']
            fr = data['fr']
            an = data['an']
            sd = data['sd']
            sr = data['sr']
            nt = data['nt']

            maximum_score = max(hp.max(), fr.max(),an.max(), sd.max(), sr.max())

            csv_results = {'Happiness': hp.tolist(), 'Fear': fr.tolist(), 'Anger': an.tolist(), 'Sadness': sd.tolist(),
                           'Surprise': sr.tolist(),'Neutrality' : nt.tolist(),'length_conversation': len(data.index),
                           'maximum_emotion_score': maximum_score}
        else:
            serializer['hasGraph'] = False

        employee = get_object_or_404(Employee, id=serializer['employee'])
        serializerEmployee = EmployeeSerializer(employee).data


        return Response(data={"analysis": serializer, "empl_info": serializerEmployee, "csv_results" : csv_results,
                              'text' : response['text_results'], 'audio' : response['audio_results'],
                              'video' : response['video_results'], 'request_id' : session.request_id, 'request': req.emotion}, status=status.HTTP_200_OK)

#create
class CreateEmployeeAPIView(APIView):
    # TODO request filter
    def post(self, request):
        if request.POST.get('email', None):
            emailField = request.POST['email']
        if request.POST.get('name', None):
            nameField = request.POST['name']
        if request.POST.get('surname', None):
            surnameField = request.POST['surname']
        if request.POST.get('birthday', None):
            birthdayField = request.POST['birthday']
        if request.POST.get('company', None):
            companyField = Company.objects.get(id=request.POST['company'])
        if request.POST.get('password', None):
            passwordField = make_password(request.POST['password'])
        stressedField = 0

        user = AppUsers.objects.create(email=emailField, password=passwordField, is_active=True, is_staff=False,
                                       is_superuser=False)
        user.save()
        employee = Employee.objects.create(birthday=birthdayField,
                                           name=nameField, surname=surnameField, company=companyField,
                                           stressed=stressedField, user=user)

        employee.save()
        return Response("Ok")
class CreateRequestAPIView(APIView):
    # TODO request filter
    def post(self, request):
        if request.POST.get('text', None):
            textField = request.POST['text']
            request = Request.objects.create(text=textField, created_at=datetime.now(), created_by=request.user.id)
            request.save()
            return Response("Ok")

        return Response("Error")

class NewRequest(APIView):
    def post(self, request):
        textField = request.data['text']
        employer = request.data['employer_id']

        Request.objects.create(text=textField, created_by=employer['id'], created_at=datetime.now())

        return Response('Ok')

#download functions
class downloadBai(APIView):
    def get(self,request):
        return download_excel('bai')
class downloadTas(APIView):
    def get(self,request):
        return download_excel('tas')
class downloadPanas(APIView):
    def get(self,request):
        return download_excel('panas')

class downloadPanas2(APIView):
    def get(self,request):
        return download_excel('panas2')
class downloadDers(APIView):
    def get(self,request):
        return download_excel('ders')
class downloadBdi(APIView):
    def get(self, request):
        return download_excel('bdi')
class downloadVas(APIView):
    def post(self, request):
        print(request.data)
        return download_excel(f'/{request.data["employee_id"]}/vas')
class downloadFirstVas(APIView):
    def post(self, request):
        print(request.data)
        return download_excel(f'{request.data["employee_id"]}/firstvas')

class VideoDownloadView(APIView):
    def get(self, request, session_id):
        session = get_object_or_404(Session, id=session_id)

        video_path = session.full_video_path

        zip_path = f'{video_path.split(".")[0]}.zip'

        # Create a zip file and add the video to it
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            zipf.write(video_path, os.path.basename(video_path))

        # Open the zip file in binary mode and read its contents
        with open(zip_path, 'rb') as zip_file:
            response = HttpResponse(zip_file.read(), content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename="video.zip"'

        # Remove the temporary zip file
        os.remove(zip_path)

        return response


def download_excel(identifier):
    path_dir = 'tmp/excel'
    path_excel = f'{path_dir}/{identifier}.xlsx'
    with open(path_excel, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename={identifier}.xlsx'
        return response



#---------------------------APPLICATION-------------------------------------
#forms
class RegistrationForm(APIView):
    permission_classes = (IsAuthenticated,)
    def post(self, request):
        employee = get_object_or_404(Employee, id=request.data['employee'])
        if employee.step != 0:
            return Response(status=status.HTTP_400_BAD_REQUEST)

        employeeData = request.data['data']
        print(employeeData)
        employeeData['step'] = 1
        name = employee.name[:2].upper()
        surname = employee.surname[:2].upper()
        if 'age' in employeeData:
            age = str(employeeData['age'])
        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)
        employeeData['username'] = f"{name}{surname}{age}"
        serializer = EmployeeGeneralSerializer(employee,data=employeeData)
        if serializer.is_valid():
            serializer.save()
            return Response(status=status.HTTP_201_CREATED)

        return Response(data=serializer.errors, status=status.HTTP_400_BAD_REQUEST)
class TasQuestionnaire(APIView):
    permission_classes = (IsAuthenticated,)
    def post(self, request):
        employee_id = request.data.get('employee', None)
        employee = get_object_or_404(Employee, id=employee_id)
        if employee.step != 1:
            return Response(data='Invalid step', status=status.HTTP_400_BAD_REQUEST)
        answers = request.data.getlist('question')
        print(answers)
        serializer = EmployeeCodeSerializer(employee)
        code = serializer.data['code']
        questions = getTasQuestions()
        createOrUpdateExcelFile(answers, 'tas', questions, code)
        employee.step = 2
        employee.save()
        return Response(status=status.HTTP_200_OK)
class BDIQuestionnaire(APIView):
    permission_classes = (IsAuthenticated,)
    def post(self, request):
        employee_id = request.data.get('employee', None)
        employee = get_object_or_404(Employee, id=employee_id)
        if employee.step != 2:
            return Response(data='Invalid step', status=status.HTTP_400_BAD_REQUEST)
        answers = request.data.getlist('question')
        print(answers)
        serializer = EmployeeCodeSerializer(employee)
        code = serializer.data['code']
        questions = getBDIQuestions()
        createOrUpdateExcelFile(answers, 'bdi', questions, code)
        employee.step = 3
        employee.save()
        return Response(status=status.HTTP_200_OK)
class BAIQuestionnaire(APIView):
    permission_classes = (IsAuthenticated,)
    def post(self, request):
        employee_id = request.data.get('employee', None)
        employee = get_object_or_404(Employee, id=employee_id)
        if employee.step != 3:
            return Response(data='Invalid step', status=status.HTTP_400_BAD_REQUEST)
        answers = request.data.getlist('question')
        print(answers)
        serializer = EmployeeCodeSerializer(employee)
        code = serializer.data['code']
        questions = getBAIQuestions()
        createOrUpdateExcelFile(answers, 'bai', questions, code)
        employee.step = 4
        employee.save()
        return Response(status=status.HTTP_200_OK)
class PANASQuestionnaire(APIView):
    permission_classes = (IsAuthenticated,)
    def post(self, request):
        employee_id = request.data.get('employee', None)
        employee = get_object_or_404(Employee, id=employee_id)
        if employee.step != 5:
            return Response(data='Invalid step', status=status.HTTP_400_BAD_REQUEST)
        answers = request.data.getlist('question')
        print(answers)
        serializer = EmployeeCodeSerializer(employee)
        code = serializer.data['code']
        questions = getPanasQuestions()
        createOrUpdateExcelFile(answers, 'panas', questions, code)
        employee.step = 6
        employee.save()
        return Response(status=status.HTTP_200_OK)
class PANAS2Questionnaire(APIView):
    permission_classes = (IsAuthenticated,)
    def post(self, request):
        employee_id = request.data.get('employee', None)
        employee = get_object_or_404(Employee, id=employee_id)
        if employee.step != 9:
            return Response(data='Invalid step', status=status.HTTP_400_BAD_REQUEST)
        answers = request.data.getlist('question')
        print(answers)
        serializer = EmployeeCodeSerializer(employee)
        code = serializer.data['code']
        questions = getPanasQuestions()
        createOrUpdateExcelFile(answers, 'panas2', questions, code)
        employee.step = 10
        employee.save()
        return Response(status=status.HTTP_200_OK)
class DERSQuestionnaire(APIView):
    permission_classes = (IsAuthenticated,)
    def post(self, request):
        employee_id = request.data.get('employee', None)
        employee = get_object_or_404(Employee, id=employee_id)
        if employee.step != 4:
            return Response(data='Invalid step', status=status.HTTP_400_BAD_REQUEST)
        answers = request.data.getlist('question')
        serializer = EmployeeCodeSerializer(employee)
        code = serializer.data['code']
        questions = getDERSQuestions()
        createOrUpdateExcelFile(answers, 'ders', questions, code)
        employee.step = 5
        employee.save()
        return Response(status=status.HTTP_200_OK)

def getTasQuestions():
    # Create dictionary for question numbers and questions
    questions = {
        1: "Sono spesso confuso circa le emozioni che provo",
        2: "Mi è difficile trovare le parole giuste per esprimere i miei sentimenti",
        3: "Provo delle sensazioni fisiche che neanche i medici capiscono",
        4: "Riesco facilmente a descrivere i miei sentimenti",
        5: "Preferisco approfondire i problemi piuttosto che descriverli semplicemente",
        6: "Quando sono sconvolto/a non so se sono triste, spaventato/a o arrabbiato/a",
        7: "Sono spesso disorientato dalle sensazioni che provo nel mio corpo",
        8: "Preferisco lasciare che le cose seguano il loro corso piuttosto che capire perché sono andate in quel modo",
        9: "Provo sentimenti che non riesco proprio ad identificare",
        10: "E’ essenziale conoscere le proprie emozioni",
        11: "Mi è difficile descrivere ciò che provo per gli altri",
        12: "Gli altri mi chiedono di parlare di più dei miei sentimenti",
        13: "Non capisco cosa stia accadendo dentro di me",
        14: "Spesso non so perché mi arrabbio",
        15: "Con le persone preferisco parlare delle cose di tutti i giorni piuttosto che delle loro emozioni",
        16: "Preferisco vedere spettacoli leggeri piuttosto che spettacoli a sfondo psicologico",
        17: "Mi è difficile rilevare i miei sentimenti più profondi anche agli amici più intimi",
        18: "Riesco a sentirmi vicino a una persona, anche se ci capita di stare in silenzio",
        19: "Trovo che l’esame dei miei sentimenti mi serve a risolvere i miei problemi personali",
        20: "Cercare significati nascosti in films o commedie distoglie dal piacere dello spettacolo"
    }
    return questions
def getPanasQuestions():
    # Create dictionary for question numbers and questions

    questions = {
             1: "Attenta/o",
             2: "Orgogliosa/o",
             3: "Turbata/o",
             4: "Entusiasta",
             5: "Forte",
             6: "Vergogna",
             7: "Irritabile",
             8: "Determinata/o",
             9: "Colpevole",
             10: "Attiva/o",
             11: "Ispirata/o",
             12: "Spaventata/o",
             13: "Nervosa/o",
             14: "Interessata/o",
             15: "Eccitata/o",
             16: "Agitata/o",
             17: "Impaurita/o",
             18: "Concentrata/o",
             19: "Ostile",
             20: "Angosciata/o"
         }


    return questions
def getBAIQuestions():
    # Create dictionary for question numbers and questions
    questions = {
        1: "Intorpidimento o formicolio",
        2: "Vampate di calore",
        3: "Gambe vacillanti",
        4: "Incapacità a rilassarsi",
        5: "Paura che qualcosa di molto brutto possa accadere",
        6: "Vertigini o sensazioni di stordimento",
        7: "Batticuore",
        8: "Umore instabile",
        9: "Agitazione in tutto il corpo",
        10: "Paura di perdere il controllo",
        11: "Respiro affannoso",
        12: "Paura di morire",
        13: "Sentirsi impauriti",
        14: "Essere terrorizzati",
        15: "Sentirsi Agitati",
        16: "Sensazione di Soffocamento",
        17: "Mani che tremano",
        18: "Dolori intestinali o di stomaco",
        19: "Sentirsi svenire",
        20: "Sentirsi arrossire",
        21: "Sentirsi sudati (non a causa del calore)"
    }
    return questions
def getBDIQuestions():
    # Create dictionary for question numbers and questions
    questions = {
        1: "Tristezza",
        2: "Pessimismo",
        3: "Fallimento",
        4: "Perdita di piacere",
        5: "Senso di colpa",
        6: "Sentimenti di punizione",
        7: "Autostima",
        8: "Autocritica",
        9: "Suicidio",
        10: "Pianto",
        11: "Agitazione",
        12: "Perdita di interessi",
        13: "Indecisione",
        14: "Senso di inutilità",
        15: "Perdita di energia",
        16: "Sonno",
        17: "Appetito",
        18: "Concentrazione",
        19: "Fatica",
        20: "Sesso"
    }
    return questions
def getDERSQuestions():
    # Create dictionary for question numbers and questions
    questions = {
        1: "Distinguo le mie emozioni",
        2: "Presto attenzione a ciò che provo",
        3: "Sento che le emozioni mi travolgono e sono fuori dal mio controllo",
        4: "Non ho idea di come mi sento",
        5: "Ho difficoltà a capire il significato dei miei sentimenti",
        6: "Sono attento ai miei sentimenti",
        7: "So esattamente quello che sto provando",
        8: "Do importanza a quello che sto provando",
        9: "Sono confusa/o rispetto a ciò che sto provando",
        10: "Quando sono turbata/o, riconosco le mie emozioni",
        11: "Quando sono turbata/o, mi arrabbio con me stessa/o perchè mi sento così",
        12: "Quando sono turbata/o, mi imbarazza sentirmi così",
        13: "Quando sono turbata/o, ho difficoltà a portare a termine il mio lavoro",
        14: "Quando sono turbata/o, perdo il controllo",
        15: "Quando sono turbata/o, credo che rimarrò così a lungo",
        16: "Quando sono turbata/o, credo che finirò col sentirmi molto depressa/o",
        17: "Quando sono turbata/o, credo che i miei sentimenti siano validi e importanti",
        18: "Quando sono turbata/o, ho difficoltà a prestare attenzione ad altre cose",
        19: "Quando sono turbata/o, mi sento fuori controllo",
        20: "Quando sono turbata/o, riesco sempre a fare le mie cose",
        21: "Quando sono turbata/o, mi vergogno di me stessa/o per il fatto sentirmi così",
        22: "Quando sono turbata/o, so che alla fine riesco a trovare un modo per sentirmi meglio",
        23: "Quando sono turbata/o, mi sento debole",
        24: "Quando sono turbata/o, mi sembra di non potere rispondere delle mie azioni",
        25: "Quando sono turbata/o, mi sento colpevole per sentirmi così",
        26: "Quando sono turbata/o, ho difficoltà a concentrarmi",
        27: "Quando sono turbata/o, ho difficoltà a controllare i miei comportamenti",
        28: "Quando sono turbata/o, credo che non ci sia niente che possa farmi stare meglio",
        29: "Quando sono turbata/o, mi irrito con me stessa/o per il fatto di sentirmi così",
        30: "Quando sono turbata/o, comincio a sentirmi molto dispiaciuta/o per me stessa/o",
        31: "Quando sono turbata/o, credo che crogiolarmi in quello stato sia tutto ciò che posso fare",
        32: "Quando sono turbata/o, perdo il controllo sui miei comportamenti",
        33: "Quando sono turbata/o, ho difficoltà a pensare ad altro",
        34: "Quando sono turbata/o, mi prendo del tempo per capire cosa sto provando",
        35: "Quando sono turbata/o, mi ci vuole molto tempo per capire cosa sto provando",
        36: "Quando sono turbata/o, sembra che le mie emozioni mi travolgano"
    }
    return questions
def createOrUpdateExcelFile(answers, identifier, questions, code):
    # Check if file exists
    path_dir = 'tmp/excel'
    path_excel = f'{path_dir}/{identifier}.xlsx'
    try:
        if not os.path.exists(path_dir):
            os.makedirs(path_dir)
        workbook = openpyxl.load_workbook(path_excel)
    except FileNotFoundError:
        # Create new workbook if file doesn't exist
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
    else:
        # Open existing worksheet
        worksheet = workbook.active

    # Add first row with Employee code
    worksheet.append(['User Code:', code])
    # Add second row with headers
    worksheet.append(['Question number', 'Question', 'User Answer'])
    # Add rows for each question
    print(f"Number of questions: {len(answers)}")
    print(answers)
    for i, a in enumerate(answers):
        print(f"Adding row for question {i + 1}")
        if identifier == "bdi" :
            row = [i+1, questions[i+1], int(a) + 1]
        else:
            row = [i+1, questions[i+1], a]

        worksheet.append(row)
    worksheet.append(['', '', ''])
    # Save workbook
    workbook.save(path_excel)

#session

class NewSession(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        name = uuid.uuid4()
        request_id = request.data.get('request_id', None)
        employee_id = request.data.get('employee_id', None)
        employee = get_object_or_404(Employee, id=employee_id)
        response = Response("Error")
        if employee.last_request_step == 3 or employee.last_request_step is None:
            if request.FILES.get('video-blob', None):
                newSession = session.createSession(employee_id, request_id)
                session_id = newSession.id
                video_file = request.FILES['video-blob']
                video_path = video.save_video(session_id, video_file, name)
                audio_path = audio.video_to_audio(session_id, name)
                text = audio.speech_to_text(session_id, name)
                newSession.full_video_path = video_path
                newSession.full_audio_path = audio_path
                newSession.text = text
                newSession.save()
                employee.step = employee.step + 1
                employee.session_no = employee.session_no + 1
                employee.last_request = request_id
                employee.last_request_step = 1
                employee.save()
                return Response("Success")
        return response

class CompleteNewRequest(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        employee_id = request.data.get('employee_id', None)
        employee = get_object_or_404(Employee, id=employee_id)
        if(employee.step != 6):
            #pick random emotion
            notCompletedRequests = Request.objects.exclude(id__in=
                                   Session.objects.filter(employee_id=employee_id).values_list('request_id', flat=True))
            if(notCompletedRequests.count() == 0):
                return Response(400)
            max = notCompletedRequests.count() - 1
            min = 0
            n = random.randint(min, max)
            serializer = RequestOnlyTextSerializer(notCompletedRequests[n])
        else:
            #trial request
            request = get_object_or_404(Request, id=1)
            serializer = RequestOnlyTextSerializer(request)
        return Response(serializer.data)
class GetQuestionnaireRequest(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        employee_id = request.data.get('employee_id', None)

        notCompletedRequest = Request.objects.exclude(id__in=
                               Questionnaire.objects.filter(employee_id=employee_id).values_list('request_id',
                            flat=True)).filter(id__in=Session.objects.filter(employee_id=employee_id).values_list('request_id',
                            flat=True)).first()
        print(notCompletedRequest)
        serializer = RequestOnlyTextSerializer(notCompletedRequest)
        print(serializer.data)
        return Response(serializer.data)

class FillInQuestionnaire(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        print(request.data)

        employee_id = request.data.get('employee_id', None)
        employee = get_object_or_404(Employee, id=employee_id)
        if employee.last_request_step == 2:
            request_id = request.data.get('request_id', None)
            request_object = get_object_or_404(Request, id=request_id)
            happiness = request.data.get('happiness', None)
            sadness = request.data.get('sadness', None)
            anger = request.data.get('anger', None)
            fear = request.data.get('fear', None)
            surprise = 0
            neutrality = request.data.get('neutrality', None)
            new_emotion = request.data.get('new_emotion', None)
            new_emotion_score = request.data.get('new_emotion_score', None)
            questionnaire = Questionnaire(employee=Employee(pk=employee_id), request=Request(pk=request_id),
                                          happiness=happiness, sadness=sadness, anger=anger, fear=fear,
                                          surprise=surprise,  new_emotion=new_emotion,
                                          new_emotion_score=new_emotion_score, neutrality=neutrality)
            questionnaire.save()

            employee.last_request_step = 3
            employee.save()
            # Check if file exists
            path_dir = f'tmp/excel/{employee_id}'
            path_excel = f'{path_dir}/vas.xlsx'
            try:
                if not os.path.exists(path_dir):
                    os.makedirs(path_dir)
                wb = openpyxl.load_workbook(path_excel)
            except FileNotFoundError:
                # Create new workbook if file doesn't exist
                wb = openpyxl.Workbook()
                ws = wb.active
                # Write header row
                ws.append(
                    [' Username', 'Request', 'Happiness', 'Sadness', 'Anger', 'Fear', 'Surprise', 'Neutrality',
                     'New Emotion', 'New Emotion Score'])
            else:
                # Open existing worksheet
                ws = wb.active

            # Write data row
            ws.append([employee.username, request_object.emotion, happiness, sadness, anger, fear, surprise, neutrality,
                       new_emotion, new_emotion_score])#

            # Save the workbook
            wb.save(path_excel)

            return Response('Ok')
        return Response(status=status.HTTP_400_BAD_REQUEST)


class retrieveQuestionnaireDataView(APIView):
    def get(self, request):
        employee_id = request.GET.get('employee_id')
        request_id = request.GET.get('request_id')

        questionnaire = get_object_or_404(Questionnaire, employee_id=employee_id, request_id=request_id)

        serializer = QuestionnaireSerializer(questionnaire)
        serialized_data = serializer.data

        return Response(serialized_data)

class RetrieveChatLogsEmployee(APIView):

    def post(self, request):
        session_id = request.data.get('chat_id', None)
        if session_id is not None:
            session_logs = SessionSerializerWithRequest(get_object_or_404(Session, pk=session_id)).data
            return Response({"logs": session_logs})

class VasQuestionnaireView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        print(request.data)

        employee_id = request.data.get('employee_id', None)
        employee = get_object_or_404(Employee, id=employee_id)
        request_id = request.data.get('request_id', None)
        request_o = get_object_or_404(Request, id=request_id)
        first = request.data.get('first', None)
        second = request.data.get('second', None)
        third = request.data.get('third', None)
        if employee.last_request_step == 1 :
            vas = Vas(employee=Employee(pk=employee_id), request=Request(pk=request_id),
                              first_question=first, second_question=second, third_question=third)
            vas.save()

            employee.last_request_step = 2
            employee.save()

            path_dir = f'tmp/excel/{employee_id}'
            path_excel = f'{path_dir}/firstvas.xlsx'
            try:
                if not os.path.exists(path_dir):
                    os.makedirs(path_dir)
                wb = openpyxl.load_workbook(path_excel)
            except FileNotFoundError:
                # Create new workbook if file doesn't exist
                wb = openpyxl.Workbook()
                ws = wb.active
                # Write header row
                ws.append(
                    [' Username', 'Request', 'Valence', 'Dominance', 'Control'])
            else:
                # Open existing worksheet
                ws = wb.active

            # Write data row
            ws.append(
                [employee.username, request_o.emotion, first, second, third])

            # Save the workbook
            wb.save(path_excel)

            return Response('Ok')

        return Response(status=status.HTTP_400_BAD_REQUEST)



#---------------------------OTHER-------------------------------------
class CreateEmployerAPIView(APIView):
    def post(self, request):
        if request.POST.get('email', None):
            emailField = request.POST['email']
        if request.POST.get('name', None):
            nameField = request.POST['name']
        if request.POST.get('surname', None):
            surnameField = request.POST['surname']
        if request.POST.get('birthday', None):
            birthdayField = request.POST['birthday']
        if request.POST.get('company', None):
            companyField = Company.objects.get(id=request.POST['company'])
        if request.POST.get('password', None):
            passwordField = make_password(request.POST['password'])

        user = AppUsers.objects.create(email=emailField, password=passwordField, is_active=True, is_staff=True,
                                       is_superuser=False)
        user.save()
        employer = Employer.objects.create(birthday=birthdayField,
                                           name=nameField, surname=surnameField, company=companyField, user=user)

        employer.save()
        return Response("Ok")

class TestVideoAnalysisAPIView(APIView):
    permission_classes = (IsAuthenticated,)
    # todo request filter
    def get(self, request):
        name = 1
        video.analyze_video(name)
        return Response("Ok")

class GetStep(APIView):
    permission_classes = (IsAuthenticated,)
    def post(self, request):
        employee = get_object_or_404(Employee, id=request['employee_id'])
        serializer = EmployeeStepSerializer(employee)

        return Response(serializer.data)


